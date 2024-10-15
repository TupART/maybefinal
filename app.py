from flask import Flask, request, render_template, send_file, redirect, url_for
import pandas as pd
from io import BytesIO
import openpyxl

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        selected_rows = request.form.getlist('selected_rows')

        # Cargar el archivo .xlsx
        df = pd.read_excel(file)

        # Validación de columnas
        if df.shape[1] < 19:
            return "El archivo debe tener al menos 19 columnas", 400

        # Crear un nuevo DataFrame para la PlantillaSTEP4.xlsx
        template_df = pd.read_excel('PlantillaSTEP4.xlsx', sheet_name=None)
        sheet = template_df['Sheet1']

        # Llenar la plantilla con los datos
        for index in selected_rows:
            idx = int(index)
            name = df.iloc[idx]['Name']
            surname = df.iloc[idx]['Surname']
            email = df.iloc[idx]['E-mail']
            pcc = df.iloc[idx]['Va a ser PCC?']
            market = df.iloc[idx]['Market']
            user_name = df.iloc[idx]['B2E User Name']

            # Rellenar columnas C, D, E, G, H, L, Q, R, V
            row = 6 + idx  # Ajustar para que empiece en la fila 7 (índice 6)

            sheet.cell(row=row, column=3, value=name)  # C
            sheet.cell(row=row, column=4, value=surname)  # D
            sheet.cell(row=row, column=5, value=email)  # E
            phone, workgroup, team, is_pcc, campaign_level = '', '', '', '', ''

            # Lógica de condiciones para llenar columnas F, G, H, L y V
            if pcc == 'Y':
                if market == 'DACH':
                    phone = '/+4940210918145 /+43122709858 /+41445295828'
                    workgroup = 'D_PCC'
                    team = 'Team_D_CCH_PCC_1'
                    is_pcc = 'Y'
                    campaign_level = 'Agent'
                elif market == 'France':
                    phone = '/+33180037979'
                    workgroup = 'F_PCC'
                    team = 'Team_F_CCH_PCC_1'
                    is_pcc = 'Y'
                    campaign_level = 'Agent'
                elif market == 'Spain':
                    phone = '/+34932952130'
                    workgroup = 'E_PCC'
                    team = 'Team_E_CCH_PCC_1'
                    is_pcc = 'Y'
                    campaign_level = 'Agent'
                elif market == 'Italy':
                    phone = '/+390109997099'
                    workgroup = 'I_PCC'
                    team = 'Team_I_CCH_PCC_1'
                    is_pcc = 'Y'
                    campaign_level = 'Agent'
            elif pcc == 'N':
                if market == 'DACH':
                    workgroup = 'D_Outbound'
                    team = 'Team_D_CCH_B2C_1'
                    is_pcc = 'N'
                    campaign_level = 'Agent'
                elif market == 'France':
                    workgroup = 'F_Outbound'
                    team = 'Team_F_CCH_B2C_1'
                    is_pcc = 'N'
                    campaign_level = 'Agent'
                elif market == 'Spain':
                    workgroup = 'E_Outbound'
                    team = 'Team_E_CCH_B2C_1'
                    is_pcc = 'N'
                    campaign_level = 'Agent'
            elif pcc == 'TL':
                if market in ['DACH', 'France', 'Spain', 'Italy']:
                    workgroup = 'D_PCC' if market in ['DACH', 'Italy'] else 'F_PCC' if market == 'France' else 'E_PCC'
                    team = f'Team_{market[0]}_CCH_PCC_1'
                    is_pcc = 'N'
                    campaign_level = 'Team Leader'
            elif pcc == 'DS':
                if market in ['DACH', 'France', 'Spain']:
                    workgroup = 'D_Outbound'
                    team = f'Team_{market[0]}_CCH_B2C_1'
                    is_pcc = 'N'
                    campaign_level = 'Agent'

            # Asignar valores a las columnas correspondientes
            sheet.cell(row=row, column=6, value=phone)  # F
            sheet.cell(row=row, column=7, value=workgroup)  # G
            sheet.cell(row=row, column=8, value=team)  # H
            sheet.cell(row=row, column=12, value=is_pcc)  # L
            sheet.cell(row=row, column=17, value=user_name)  # Q
            sheet.cell(row=row, column=18, value=user_name)  # R
            sheet.cell(row=row, column=22, value=campaign_level)  # V

        # Guardar el archivo en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet.to_excel(writer, index=False)

        output.seek(0)
        return send_file(output, attachment_filename='PlantillaSTEP4.xlsx', as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
